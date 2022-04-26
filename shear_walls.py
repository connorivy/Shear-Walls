from shear_wall_classes import *
from openpyxl import load_workbook
import argparse

def main():
    project = Project()

    path = "P:\\\\Structural\\2021\\2115 - Dallas\\2115.147 Toll Brothers Addison Road\\5 - Project Documents\B - Calculations\CI\Wood Calcs\Shear Walls\\"
    #load excel file
    workbook = load_workbook(filename=path + "toll_brothers_shear_wall_input.xlsx", data_only=True)

    ws = workbook['input']

    prev_shear_wall_letter = ws['A2'].value
    wall_args = []
    blank_count = 0
    current_row = 0

    parser = argparse.ArgumentParser()
    parser.add_argument('--start', type=str, required=True)
    parser.add_argument('--end', type=str, required=True)
    args = parser.parse_args()

    start_letters_no_nums = ''.join([i for i in args.start if not i.isdigit()])
    rows_to_delete = []
    delete_spaces = True

    for row in ws.iter_rows(min_row=2, max_col=8, max_row=1500, values_only=True):
        current_row += 1
        # quit the for loop once three consecutive blanks have been found
        if row[0] == None:
            if delete_spaces:
                rows_to_delete.append(current_row+1)
            blank_count += 1
            if blank_count > 3:
                current_row -= 2
                break
            continue
        else:
            blank_count = 0

        current_letters_no_nums = ''.join([i for i in row[0] if not i.isdigit()])
        if len(current_letters_no_nums) < len(start_letters_no_nums):
            rows_to_delete.append(current_row+1)
            continue
        elif row[0] < args.start:
            rows_to_delete.append(current_row+1)
            continue
        elif row[0] > args.end:
            rows_to_delete.append(current_row+1)
            delete_spaces = True
            continue
        else:
            delete_spaces = False
            
        if row[0] != prev_shear_wall_letter:
            write_values_to_excel(ws, project, wall_args, current_row, prev_shear_wall_letter)
            wall_args = []
            prev_shear_wall_letter = row[0]

        if row[6]:
            wall_args.append(*[row[1:7]])
        else:
            wall_args.append(*[row[1:6]])

    write_values_to_excel(ws, project, wall_args, current_row, prev_shear_wall_letter)
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
    workbook.save(path + "shear_wall_output.xlsx")

def write_values_to_excel(ws, project, wall_args, row, wall_location):
    num_rows = len(wall_args)
    stacked_shear_wall = StackedShearWall(project, *wall_args)

    for i in range(num_rows):
        schedule_entry = project.shear_wall_schedule.get_shear_wall(stacked_shear_wall.shear_force[i], max(0,stacked_shear_wall.chord_forces[i] / 1000), wall_location)
        ws.cell(row = row - num_rows + i, column = 13, value=stacked_shear_wall.shear_force[i] / 1000)
        ws.cell(row = row - num_rows + i, column = 14, value=stacked_shear_wall.overturning_moment[i] / 1000)
        ws.cell(row = row - num_rows + i, column = 15, value=stacked_shear_wall.resisting_moment[i] / 1000)
        ws.cell(row = row - num_rows + i, column = 16, value=stacked_shear_wall.total_moment[i] / 1000)
        ws.cell(row = row - num_rows + i, column = 17, value=max(0,stacked_shear_wall.chord_forces[i] / 1000))
        try:
            ws.cell(row = row - num_rows + i, column = 18, value=schedule_entry.name)
        except:
            ws.cell(row = row - num_rows + i, column = 18, value=f'ERROR AT WALL {wall_location} - NO VALUE HIGH ENOUGH')


main()